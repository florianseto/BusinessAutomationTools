"""excel_writer.py の単体テスト（T030～T038）。"""

import pandas as pd
import pytest
from openpyxl import load_workbook

from excel_writer import write_excel


def sample_frame():
    return pd.DataFrame({"社員番号": ["001", "002"], "氏名": ["山田", "佐藤"]})


def test_t030_write_excel(tmp_path):
    output = tmp_path / "output.xlsx"
    write_excel(sample_frame(), str(output))
    actual = pd.read_excel(output, dtype={"社員番号": str})
    pd.testing.assert_frame_equal(actual, sample_frame())


def test_t031_default_sheet_name(tmp_path):
    output = tmp_path / "output.xlsx"
    write_excel(sample_frame(), str(output))
    assert load_workbook(output, read_only=True).sheetnames == ["加工済みデータ"]


def test_t032_custom_sheet_name(tmp_path):
    output = tmp_path / "output.xlsx"
    write_excel(sample_frame(), str(output), sheet_name="勤務実績")
    assert load_workbook(output, read_only=True).sheetnames == ["勤務実績"]


def test_t033_create_nested_output_directory(tmp_path):
    output = tmp_path / "level1" / "level2" / "output.xlsx"
    write_excel(sample_frame(), str(output))
    assert output.is_file()


def test_t034_do_not_write_dataframe_index(tmp_path):
    output = tmp_path / "output.xlsx"
    source = pd.DataFrame({"A": [10, 20]}, index=[100, 200])
    write_excel(source, str(output))
    actual = pd.read_excel(output)
    assert actual.columns.tolist() == ["A"]
    assert actual["A"].tolist() == [10, 20]


def test_t035_write_empty_dataframe(tmp_path):
    output = tmp_path / "output.xlsx"
    write_excel(pd.DataFrame(columns=["A", "B"]), str(output))
    actual = pd.read_excel(output)
    assert actual.empty
    assert actual.columns.tolist() == ["A", "B"]


def test_t036_overwrite_existing_file(tmp_path):
    output = tmp_path / "output.xlsx"
    write_excel(pd.DataFrame({"old": [1]}), str(output))
    expected = pd.DataFrame({"new": [2]})
    write_excel(expected, str(output))
    pd.testing.assert_frame_equal(pd.read_excel(output), expected)


def test_t037_invalid_sheet_name(tmp_path):
    with pytest.raises(ValueError):
        write_excel(sample_frame(), str(tmp_path / "output.xlsx"), sheet_name="A" * 32)


def test_t038_unwritable_parent_path(tmp_path):
    parent_file = tmp_path / "not_a_directory"
    parent_file.write_text("file", encoding="utf-8")
    with pytest.raises(OSError):
        write_excel(sample_frame(), str(parent_file / "output.xlsx"))

